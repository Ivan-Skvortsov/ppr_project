from datetime import date

from openpyxl import load_workbook
from tqdm import tqdm

from reports.models import (EquipmentType, Facility, MaintenanceCategory,
                            MaintenanceType, Schedule)


def import_objects_from_xls(filename, category_name):
    """Imports objects from xlsx file."""
    try:
        category = MaintenanceCategory.objects.get(category_name=category_name)
    except Exception as err:
        print(f'Неверная категория {category_name}: {err}')
        raise
    wb = load_workbook(filename=filename, read_only=True, data_only=True)
    worksheet = wb.active
    for row in tqdm(worksheet.iter_rows(min_row=2), total=worksheet.max_row - 1):
        facility_name = row[0].value
        if not facility_name:
            return
        eqipment_type_name = row[1].value

        facility = Facility.objects.get_or_create(
            facility_name=facility_name,
            maintenance_category=category
        )[0]

        EquipmentType.objects.get_or_create(
            facility=facility,
            eqipment_type_name=eqipment_type_name,
            maintenance_category=category
        )


def import_schedule_from_xls(filename, category_name):
    """Import schedules from xlsx file."""
    wb = load_workbook(filename=filename, read_only=True, data_only=True)
    worksheet = wb.active
    for row in tqdm(worksheet.iter_rows(min_row=2), total=worksheet.max_row - 1):
        category = MaintenanceCategory.objects.get(category_name=category_name)
        facility_name = row[0].value
        if not facility_name:
            return
        facility = Facility.objects.get(
            facility_name=facility_name,
            maintenance_category=category
        )
        equipment_type = EquipmentType.objects.get(
            facility=facility,
            maintenance_category=category,
            eqipment_type_name=row[1].value
        )
        for i in range(2, 14):
            if row[i].value:
                m_type = MaintenanceType.objects.get_or_create(
                    m_type=row[i].value.strip()
                )[0]
                date_scheduled = date(2024, i-1, 1)
                Schedule.objects.create(
                    equipment_type=equipment_type,
                    maintenance_type=m_type,
                    date_sheduled=date_scheduled
                )


def execute_import_schedules_and_objects(filename):
    category_name = input('Введите категорию: ')
    print('Creating objects: \n')
    import_objects_from_xls(filename, category_name)
    print('Creating schedules: \n')
    import_schedule_from_xls(filename, category_name)
    print('Finished!')
