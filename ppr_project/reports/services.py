from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime

from django.db.models import Q, QuerySet
from django.template.defaultfilters import date as _date

from openpyxl import Workbook, worksheet
from openpyxl.styles import Alignment, Border, Font, Side
from simple_history.utils import bulk_update_with_history

from reports.models import Schedule


class XlsxReportGenerator:

    FONT_BOLD = Font(name='Times New Roman', size=12, bold=True)
    FONT_REGULAR = Font(name='Times New Roman', size=12)

    ALIGNMENT_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGNMENT_TOP_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
    ALIGNMENT_TOP_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)

    BORDER_ALL_SIDES = Border(
        bottom=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000')
    )
    BORDER_BOTTOM = Border(bottom=Side(border_style='thin', color='FF000000'))

    def __init__(self, date_from: date, date_to: date, report_type: str):
        self.date_from = date_from
        self.date_to = date_to
        self.report_type = report_type

    def _get_queryset(self) -> QuerySet:
        qs_filter = {
            'ppr': Q(maintenance_type__m_type__icontains='ТО') & Q(date_completed__isnull=False),
            'ppz': Q(maintenance_type__m_type__icontains='Проверка') & Q(date_completed__isnull=False),
            'asps': Q(equipment_type__eqipment_type_name__icontains='АСПС') & Q(date_completed__isnull=False),
            'uncompletable': Q(uncompleted__isnull=False),
        }
        qs = Schedule.objects.filter(
            date_sheduled__gte=self.date_from,
            date_sheduled__lte=self.date_to
        ).order_by('date_completed', 'equipment_type__facility')
        return qs.filter(qs_filter[self.report_type])

    def _render_ppr_ppz_report_header(self, ws: worksheet) -> None:
        report_header = {
            'ppr': 'ППР оборудования АСУ, А и ТМ',
            'ppz': 'проверок защит'
        }
        date_from = datetime.strptime(self.date_from, '%Y-%m-%d')
        date_to = datetime.strptime(self.date_to, '%Y-%m-%d')
        ws['G1'] = 'Утверждаю'
        ws['A7'] = f'Протокол проведения {report_header[self.report_type]}'
        ws['A8'] = 'объектов КС-45 Усинская'
        ws['A9'] = (f'за период с {_date(date_from, "d E")} по '
                    f'{_date(date_to, "d E Y")} г.')
        ws['A11'] = ('Основание проведения работ: График проведения ППР '
                     'оборудования АСУ, А и ТМ КС-45 Усинская на 2022 год.')
        ws['A12'] = ''
        ws.append(['№ п/п',
                   'Объект',
                   'Наименование объекта/системы',
                   'Тип ТО',
                   'Результат',
                   'Дата',
                   'Ответственные лица'])

    def _render_employees_string(self, entry: Schedule) -> str:
        employees = [
            f'{e.position}\n{e.name}\n\n' for e in (entry.employee1, entry.employee2, entry.employee3)
            if e is not None and e.position != 'Приборист'
        ]
        return ''.join(employees)

    def _render_ppr_ppz_report_body(self, ws: worksheet) -> None:
        queryset = self._get_queryset()
        if queryset.count() == 0:
            ws.append(
                ['', '', 'За выбранный период завершенных работ не найдено']
            )
            return
        for i, work in enumerate(queryset):
            employees = self._render_employees_string(work)
            date_completed = _date(work.date_completed, 'd.m.Y')
            ws.append(
                [i + 1,
                 work.equipment_type.facility.facility_name,
                 work.equipment_type.eqipment_type_name,
                 work.maintenance_type.m_type,
                 'Выполнено.\nЗамечаний нет.',
                 date_completed,
                 employees]
            )

    def _post_process_ppr_ppz_report(self, ws: worksheet) -> None:
        """Merge cells with same data in cols 2, 6, 7 (Объект, Дата, ФИО)."""
        start_row = 14
        end_row = ws.max_row
        facility = None
        employees = None
        merge_count = 0
        row_iterator = ws.iter_rows(min_row=start_row, max_row=end_row)
        for row, cells in enumerate(row_iterator, start=start_row):
            if cells[1].value == facility and cells[6].value == employees:
                merge_count += 1
            else:
                if merge_count > 0:
                    ws.merge_cells(start_column=2, end_column=2, start_row=row - merge_count - 1, end_row=row - 1)
                    ws.merge_cells(start_column=6, end_column=6, start_row=row - merge_count - 1, end_row=row - 1)
                    ws.merge_cells(start_column=7, end_column=7, start_row=row - merge_count - 1, end_row=row - 1)
                facility = cells[1].value
                employees = cells[6].value
                merge_count = 0
        if merge_count > 0:
            ws.merge_cells(start_column=2, end_column=2, start_row=row - merge_count, end_row=row)
            ws.merge_cells(start_column=6, end_column=6, start_row=row - merge_count, end_row=row)
            ws.merge_cells(start_column=7, end_column=7, start_row=row - merge_count, end_row=row)

    def _style_ppr_ppz_report(self, ws: worksheet):

        # worksheet styling
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 19
        ws.column_dimensions['C'].width = 33
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 19
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 26
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0
        ws.page_margins.footer = 0

        ws.merge_cells(start_row=7, end_row=7, start_column=1, end_column=7)
        ws.merge_cells(start_row=8, end_row=8, start_column=1, end_column=7)
        ws.merge_cells(start_row=9, end_row=9, start_column=1, end_column=7)
        ws.merge_cells(start_row=11, end_row=11, start_column=1, end_column=7)

        for row in ws.iter_rows(min_row=1, max_row=13):
            for cell in row:
                cell.font = self.FONT_BOLD
                cell.alignment = self.ALIGNMENT_CENTER
        for cell in ws[13]:
            cell.border = self.BORDER_ALL_SIDES

        ws['G2'].border = ws['G3'].border = ws['G4'].border = self.BORDER_BOTTOM
        ws['A11'].alignment = Alignment(horizontal='left')
        ws['A11'].font = self.FONT_REGULAR

        for row in ws.iter_rows(min_row=14, max_row=ws.max_row):
            for cell in row:
                cell.font = self.FONT_REGULAR
                cell.border = self.BORDER_ALL_SIDES
                if cell.column in [1, 4, 5, 6]:
                    cell.alignment = self.ALIGNMENT_TOP_CENTER
                else:
                    cell.alignment = self.ALIGNMENT_TOP_LEFT

    def _render_ppr_ppz_report(self) -> Workbook:
        wb, ws = self._create_workbook_with_empty_worksheet()
        self._render_ppr_ppz_report_header(ws)
        self._render_ppr_ppz_report_body(ws)
        self._post_process_ppr_ppz_report(ws)
        self._style_ppr_ppz_report(ws)
        return wb

    def _create_workbook_with_empty_worksheet(self):
        wb = Workbook()
        del wb['Sheet']
        ws = wb.create_sheet(self.report_type)
        return wb, ws

    def _render_asps_report(self) -> Workbook:
        wb, ws = self._create_workbook_with_empty_worksheet()
        queryset = self._get_queryset()
        ws.append(
            ['Дата проверки',
             'Местонахождение системы противопожарной защиты',
             'Вид мероприятия по эксплуатации систем противопожарной защиты',
             'Наличие/отсутствие замечаний',
             'Подтверждающий документ (акт, протокол)',
             'Должность,Ф.И.О проводившего проверку']
        )
        ws.append([1, 2, 3, 4, 5, 6])
        if queryset.count() == 0:
            ws.append(
                ['', '', 'За выбранный период завершенных работ не найдено']
            )
            return
        for work in queryset:
            employees = self._render_employees_string(work)
            date_completed = _date(work.date_completed, 'd.m.Y')
            work_type_and_name = f'{work.maintenance_type.m_type}: {work.equipment_type.eqipment_type_name}'
            ws.append(
                [date_completed,
                 work.equipment_type.facility.facility_name,
                 work_type_and_name,
                 'Замечаний нет',
                 f'Запись в журнале от {date_completed}',
                 employees.replace('\n', ' ')]
            )

        # styling
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 27
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 55

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.font = self.FONT_REGULAR
                cell.border = self.BORDER_ALL_SIDES
                cell.alignment = self.ALIGNMENT_CENTER
        return wb

    def _render_uncompletable_report(self) -> Workbook:
        wb, ws = self._create_workbook_with_empty_worksheet()
        queryset = self._get_queryset()
        today_date = _date(date.today(), 'd E Y')
        date_from = _date(datetime.strptime(self.date_from, '%Y-%m-%d'), 'd.m.Y')
        date_to = _date(datetime.strptime(self.date_to, '%Y-%m-%d'), 'd.m.Y')
        ws.append([today_date, '', '', '', '', 'КС-45 Усинская'])
        ws['A3'] = 'Протокол\nвыполнения ППР службой АСУ, А и ТМ'
        ws['A5'] = (
            f'          Настоящий протокол составлен о том, что в период с {date_from} по {date_to}, согласно графика '
            'ППР службы АСУ, А и ТМ КС-45 «Усинская» на КЦ-1,2 ПБ КС-45 Усинская проведены работы по техническому '
            'обслуживанию оборудования автоматизации. Работы выполнены в полном объеме, с надлежащим качеством.'
        )
        if queryset:
            ws['A7'] = '          Не проведены работы:'
            ws.append(['№ п/п', 'Категория', 'Объект', 'Оборудование', 'Вид ТО', 'Причина невыполнения'])
            for i, work in enumerate(queryset):
                ws.append([
                    i + 1,
                    work.equipment_type.facility.maintenance_category.category_name,
                    work.equipment_type.facility.facility_name,
                    work.equipment_type.eqipment_type_name,
                    work.maintenance_type.m_type,
                    work.uncompleted.reason
                ])
        ws.append([''])
        ws.append(['Инженер АСУ, А и ТМ КС-45 Усинская'])
        ws.append([''])
        ws.append(['Приборист АСУ, А и ТМ КС-45 Усинская'])

        # styling
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 36
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 16
        ws.row_dimensions[3].height = 31
        ws.row_dimensions[5].height = 62
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=3)
        ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=6)
        ws.merge_cells(start_row=5, end_row=5, start_column=1, end_column=6)
        ws.merge_cells(start_row=7, end_row=7, start_column=1, end_column=6)
        ws.merge_cells(start_row=ws.max_row - 2, end_row=ws.max_row - 2, start_column=1, end_column=6)
        ws.merge_cells(start_row=ws.max_row, end_row=ws.max_row, start_column=1, end_column=6)
        for row in ws.iter_rows(1, ws.max_row):
            for cell in row:
                cell.font = self.FONT_REGULAR
                cell.alignment = self.ALIGNMENT_TOP_LEFT
                if queryset and (8 <= cell.row <= ws.max_row - 4):
                    cell.border = self.BORDER_ALL_SIDES
        ws['A3'].alignment = self.ALIGNMENT_CENTER
        if queryset:
            for cell in ws[8]:
                cell.font = self.FONT_BOLD
                cell.alignment = self.ALIGNMENT_CENTER
        return wb

    def render_report(self):
        """Entry point for xlsx report generation."""
        render_report_methods = {
            'ppr': self._render_ppr_ppz_report,
            'ppz': self._render_ppr_ppz_report,
            'asps': self._render_asps_report,
            'uncompletable': self._render_uncompletable_report
        }
        return render_report_methods[self.report_type]()


def distribute_next_month_works_by_dates():
    """Distributes next month schedules by dates."""
    current_date = date.today()
    current_year = current_date.year
    next_month = current_date.month + 1
    last_day_of_month = monthrange(current_year, next_month)[1]
    qs = (Schedule.objects.select_related('equipment_type__facility')
                          .filter(date_sheduled__month=next_month)
                          .order_by('equipment_type__facility'))
    regouped_schedules = defaultdict(list)
    for schedule in qs:
        regouped_schedules[schedule.equipment_type.facility].append(schedule)
    current_day = 1
    for group in regouped_schedules:
        for schedule in regouped_schedules[group]:
            schedule.date_sheduled = date(
                year=current_year, month=next_month, day=current_day
            )
            schedule._change_reason = 'Distributed next month shedules by day'
        current_day += 1
        if current_day > last_day_of_month:
            current_day = 1
    bulk_update_with_history(
        qs, Schedule, ['date_sheduled'], batch_size=500
    )


def get_next_month_plans():
    current_date = date.today()
    next_month = current_date.month + 1
    qs = (Schedule.objects.select_related('equipment_type__facility',
                                          'maintenance_type',
                                          'equipment_type')
                          .order_by('date_sheduled',
                                    'equipment_type__facility')
                          .filter(date_sheduled__month=next_month))
    regrouped_schedules = defaultdict(list)
    for schedule in qs:
        regrouped_schedules[schedule.date_sheduled].append(schedule)
    wb = Workbook()
    ws = wb.active
    for schedule_date in regrouped_schedules:
        formatted_date = _date(schedule_date, 'd E Y')
        ws.append([formatted_date])
        for schedule in regrouped_schedules[schedule_date]:
            ws.append([
                schedule.equipment_type.facility.facility_name,
                schedule.equipment_type.eqipment_type_name,
                schedule.maintenance_type.m_type
            ])
        ws.append([''])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    return wb
